from excel.worksheet import *
from openpyxl.utils.dataframe import dataframe_to_rows


def result(worksheet, dataframe, problem):
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        worksheet.append(row)

    border_rows(worksheet, 'thin', 'FF000000',
                top_rows=[1, 2],
                bottom_rows=[worksheet.max_row])
    border_columns(worksheet,'thin', 'FF000000',
                   left_columns=[1, 2, 3, 4, 5, 11, 15, 21, 25, 26],
                   right_columns=[worksheet.max_column])


    number_format_columns(worksheet, '0.000', columns=[4, 13, 21, 23, 25])
    number_format_columns(worksheet, '0.00',  columns=[11])
    number_format_columns(worksheet, '0',     columns=[2, 3, 5, 7, 9, 15, 17, 19])
    if problem["cost_style"] == "percent":
        number_format_columns(worksheet, '0.000%', columns=[5, 15])
    number_format_rows(worksheet, 'General', rows=[1])

    columns = [1, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24]
    alignment_columns(worksheet, px.styles.Alignment(wrap_text=False, horizontal='center', vertical='center'),
                      columns=columns)
    alignment_columns(worksheet, px.styles.Alignment(wrap_text=False, horizontal='right', vertical='center'),
                      columns=[column for column in range(1, worksheet.max_column + 1) if columns])
    alignment_rows(worksheet, px.styles.Alignment(wrap_text=False, horizontal='right', vertical='center'),
                   rows=[1])
    width_auto(worksheet)

