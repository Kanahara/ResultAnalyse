import pandas as pd
import sys
import colorsys
import random
from openpyxl.utils.dataframe import dataframe_to_rows
from excel.worksheet import *
import operator
import openpyxl as px


def comparison_result(worksheet, algorithms, problem):
    dataframes = [algorithm.result_dataframe() for algorithm in algorithms]
    columns = ["", "", "", ""]
    dataframe_list = [dataframes[0].iloc[:, [0, 1, 2, 3]]]
    for basis_column, dataframe in enumerate(dataframes):
        dataframe_list.append(dataframe.iloc[:, [4, 5, 6, 9, 10]])
        columns.extend([str(basis_column), "", "", "", ""])

    for basis_column, dataframe in enumerate(dataframes):
        dataframe_list.append(dataframe.iloc[:, [20]])
        columns.append(str(basis_column))

    worksheet.append(columns)
    for row in dataframe_to_rows(pd.concat(dataframe_list, axis=1), index=False, header=True):
        worksheet.append(row)

    general_columns = [1]
    int_columns = [2, 3]
    float_columns = [4]
    percent_columns = []
    border_left_columns = [1, 2, 3, 4]
    dataframe_columns = []

    r = random.random()
    for basis_column in range(5, 5 * (len(dataframes) + 1), 5):
        if problem["cost_style"] == "percent":
            percent_columns.extend([basis_column, basis_column + 4])
        else:
            int_columns.extend([basis_column, basis_column + 4])
        general_columns.extend([basis_column + 1, basis_column + 3])
        int_columns.extend([basis_column + 2])
        border_left_columns.extend([basis_column, basis_column + 4, 4 + 5 * len(dataframes) + basis_column // 5])
        float_columns.extend([4 + 5 * len(dataframes) + basis_column // 5])
        color = colorsys.hsv_to_rgb(r, 0.1, 0.9)
        red = format(round(color[0] * 255), '02x')
        green = format(round(color[1] * 255), '02x')
        blue = format(round(color[2] * 255), '02x')
        r += random.uniform(0.15, 0.85)
        fill_columns(worksheet, px.styles.PatternFill(patternType='solid', fgColor=red+green+blue),
                     columns=[basis_column, basis_column + 1, basis_column + 2, basis_column + 3, basis_column + 4, 4 + 5 * len(dataframes) + basis_column // 5])

    number_format_columns(worksheet, "General", columns=general_columns)
    number_format_columns(worksheet, "0",       columns=int_columns)
    number_format_columns(worksheet, "0.000",   columns=float_columns)
    number_format_columns(worksheet, "0.000%",  columns=percent_columns)
    number_format_rows(worksheet, "General", rows=[1, 2])
    border_columns(worksheet,'thin', 'FF000000', left_columns=border_left_columns, right_columns=[worksheet.max_column])
    border_rows(worksheet,'thin', 'FF000000', top_rows=[1, 2], bottom_rows=[2, worksheet.max_row])

    alignment_columns(worksheet, px.styles.Alignment(wrap_text=False, horizontal='center', vertical='center'),
                      columns=general_columns)
    alignment_columns(worksheet, px.styles.Alignment(wrap_text=False, horizontal='right', vertical='center'),
                      columns=int_columns + float_columns + percent_columns)
    alignment_rows(worksheet, px.styles.Alignment(wrap_text=False, horizontal='center', vertical='center'),
                   rows=[1, 2])
    width_auto(worksheet)

    for row in range(3, worksheet.max_row + 1):
        best_best_cost = sys.float_info.min if problem["type"] == "maximization" else sys.float_info.max
        best_avg_cost = sys.float_info.min if problem["type"] == "maximization" else sys.float_info.max
        max_trial = sys.float_info.min
        min_time = sys.float_info.max

        bold_best_cost_columns = []
        bold_avg_cost_columns = []
        bold_trial_columns = []
        bold_time_columns = []
        for basis_column in range(5, 5 * (len(dataframes) + 1), 5):
            best_cost_column = basis_column
            best_cost = worksheet.cell(row=row, column=basis_column).value
            trial_column = basis_column + 2
            trial = worksheet.cell(row=row, column=trial_column).value
            avg_cost_column = basis_column + 4
            avg_cost = worksheet.cell(row=row, column=avg_cost_column).value
            time_column = 4 + 5 * len(dataframes) + basis_column // 5
            time = worksheet.cell(row=row, column=time_column).value
            comp = operator.gt if problem["type"] == "maximization" else operator.lt
            if comp(best_cost, best_best_cost):
                best_best_cost = best_cost
                bold_best_cost_columns = [best_cost_column]
                max_trial = trial
                bold_trial_columns = [trial_column]
            elif best_cost == best_best_cost:
                bold_best_cost_columns.append(best_cost_column)
                if trial > max_trial:
                    max_trial = trial
                    bold_trial_columns = [trial_column]
                elif trial == max_trial:
                    bold_trial_columns.append(trial_column)

            if comp(avg_cost, best_avg_cost):
                best_avg_cost = avg_cost
                bold_avg_cost_columns = [avg_cost_column]
            elif avg_cost == best_avg_cost:
                bold_avg_cost_columns.append(avg_cost_column)

            if time < min_time:
                min_time = time
                bold_time_columns = [time_column]
            elif time == min_time:
                bold_time_columns.append(time_column)

        underline_time_columns = []
        if len(bold_best_cost_columns) == len(bold_trial_columns) and len(bold_best_cost_columns) == len(bold_avg_cost_columns):
            best_cost_min_time = sys.float_info.max
            for basis_column in range(5, 5 * (len(dataframes) + 1), 5):
                time_column = 4 + 5 * len(dataframes) + basis_column // 5
                time = worksheet.cell(row=row, column=time_column).value
                if time < best_cost_min_time:
                    best_cost_min_time = time
                    underline_time_columns = [time_column]
                elif time == best_cost_min_time:
                    underline_time_columns.append(time_column)

        for column in bold_best_cost_columns + bold_avg_cost_columns + bold_trial_columns + bold_time_columns:
            underline = 'single' if column in underline_time_columns else None
            worksheet.cell(row=row, column=column).font = px.styles.fonts.Font(b=True, u=underline)

    add_column = worksheet.max_column + 2
    for index, algorithm in enumerate(algorithms):
        worksheet.cell(row=index + 1, column=add_column).value = str(index)
        worksheet.cell(row=index + 1, column=add_column + 1).value = algorithm.name()
