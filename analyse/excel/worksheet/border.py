import openpyxl as px


def border_rows(worksheet, style, color, top_rows, bottom_rows):
    for column in range(1, worksheet.max_column + 1):
        for row in top_rows:
            top = px.styles.Side(style=style, color=color)
            bottom = worksheet.cell(row=row, column=column).border.bottom
            left = worksheet.cell(row=row, column=column).border.left
            right = worksheet.cell(row=row, column=column).border.right
            worksheet.cell(row=row, column=column).border = px.styles.Border(top=top, bottom=bottom, left=left, right=right)
        for row in bottom_rows:
            top = worksheet.cell(row=row, column=column).border.top
            bottom = px.styles.Side(style=style, color=color)
            left = worksheet.cell(row=row, column=column).border.left
            right = worksheet.cell(row=row, column=column).border.right
            worksheet.cell(row=row, column=column).border = px.styles.Border(top=top, bottom=bottom, left=left, right=right)


def border_columns(worksheet, style, color, left_columns, right_columns):
    for row in range(1, worksheet.max_row + 1):
        for column in left_columns:
            top = worksheet.cell(row=row, column=column).border.top
            bottom = worksheet.cell(row=row, column=column).border.bottom
            left = px.styles.Side(style=style, color=color)
            right = worksheet.cell(row=row, column=column).border.right
            worksheet.cell(row=row, column=column).border = px.styles.Border(top=top, bottom=bottom, left=left, right=right)
        for column in right_columns:
            top = worksheet.cell(row=row, column=column).border.top
            bottom = worksheet.cell(row=row, column=column).border.bottom
            left = worksheet.cell(row=row, column=column).border.left
            right = px.styles.Side(style=style, color=color)
            worksheet.cell(row=row, column=column).border = px.styles.Border(top=top, bottom=bottom, left=left, right=right)
