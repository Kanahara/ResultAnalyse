import math
import numpy as np
import openpyxl as px


def width_auto(worksheet):
    for column in range(1, worksheet.max_column + 1):
        max_len = 0
        for row in range(1, worksheet.max_row + 1):
            str_len = 0
            if type(worksheet.cell(row=row, column=column).value) is float and np.isnan(worksheet.cell(row=row, column=column).value):
                pass
            elif worksheet.cell(row=row, column=column).number_format == '0':
                str_len = 1.0 + __number_of_digits__(int(worksheet.cell(row=row, column=column).value))
            elif worksheet.cell(row=row, column=column).number_format == '0.0':
                str_len = 2.5 + __number_of_digits__(int(worksheet.cell(row=row, column=column).value))
            elif worksheet.cell(row=row, column=column).number_format == '0.00':
                str_len = 3.5 + __number_of_digits__(int(worksheet.cell(row=row, column=column).value))
            elif worksheet.cell(row=row, column=column).number_format == '0.000':
                str_len = 4.5 + __number_of_digits__(int(worksheet.cell(row=row, column=column).value))
            elif worksheet.cell(row=row, column=column).number_format == '0.000%':
                str_len = 5.5 + __number_of_digits__(int(worksheet.cell(row=row, column=column).value))
            else:
                str_len = len(str(worksheet.cell(row=row, column=column).value))
            max_len = max(max_len, str_len)
        worksheet.column_dimensions[px.utils.get_column_letter(column)].width = max_len * 1.2


def __number_of_digits__(num):
    if num == 0:
        return 1
    else:
        return math.log10(num) + 1